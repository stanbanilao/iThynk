import io
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pypdf import PdfReader


DEFAULT_MAPPINGS = [
    ("Consumer Name", ["consumer name", "first name", "forenames"]),
    ("Consumer Surname", ["consumer surname", "surname", "last name"]),
    ("ID Number", ["identity number", "id number", "rsa id"]),
    ("Contact Number", ["contact number", "cellphone", "mobile number", "telephone"]),
    ("Employer", ["employer details", "employer", "company name"]),
    ("Status", ["applicant status", "consumer status", "status"]),
    ("Status Date", ["status date", "date of status"]),
    ("iDocs Reference", ["reference number", "reference no", "reference"]),
]


@dataclass
class ExtractedField:
    field_name: str
    value: str
    pdf_label: str
    page_number: int
    confidence: str
    review_required: str

    def as_preview(self):
        return {
            "field": self.field_name,
            "value": self.value,
            "pdf_label": self.pdf_label,
            "page": self.page_number,
            "confidence": self.confidence,
            "review_required": self.review_required,
        }


def normalise(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def mask_id(value):
    digits = re.sub(r"\D", "", str(value or ""))
    return "*" * max(0, len(digits) - 4) + digits[-4:]


def extract_pdf_text(pdf_bytes):
    if not pdf_bytes:
        raise ValueError("The iDocs PDF preview returned no data")
    reader = PdfReader(io.BytesIO(pdf_bytes))
    pages = [(index + 1, page.extract_text() or "") for index, page in enumerate(reader.pages)]
    if not pages:
        raise ValueError("The iDocs PDF has no readable pages")
    return pages


def extract_fields(pages, mappings=None):
    mappings = mappings or DEFAULT_MAPPINGS
    results = []
    for field_name, labels in mappings:
        found = None
        for page_number, text in pages:
            lines = [normalise(line) for line in text.splitlines() if normalise(line)]
            for index, line in enumerate(lines):
                for label in labels:
                    match = re.search(rf"(?i)\b{re.escape(label)}\b\s*[:\-]?\s*(.*)$", line)
                    if not match:
                        continue
                    value = normalise(match.group(1))
                    if not value and index + 1 < len(lines):
                        value = normalise(lines[index + 1])
                    if value:
                        found = ExtractedField(field_name, value, label, page_number, "Label match", "No")
                        break
                if found:
                    break
            if found:
                break
        results.append(found or ExtractedField(field_name, "", "", 0, "Not found", "Yes"))
    return results


def verify_document_id(submitted_id, extracted):
    expected = re.sub(r"\D", "", str(submitted_id or ""))
    actual = ""
    for item in extracted:
        if item.field_name == "ID Number":
            actual = re.sub(r"\D", "", item.value)
            break
    if not actual:
        return "Not verified", "The document ID could not be extracted"
    if expected == actual:
        return "Verified", ""
    return "Mismatch", "The submitted ID does not match the document ID"


def read_pdf_preview(pdf_bytes, submitted_id):
    pages = extract_pdf_text(pdf_bytes)
    extracted = extract_fields(pages)
    verification, note = verify_document_id(submitted_id, extracted)
    return pages, extracted, verification, note


def write_calibration_workbook(output_path, submitted_id, source_name, pages, extracted, method="Selectable PDF text"):
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    verification, verification_note = verify_document_id(submitted_id, extracted)
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"
    headers = ["Field", "Extracted Value", "PDF Label", "Page", "Confidence", "Review Required"]
    ws.append(headers)
    for item in extracted:
        ws.append([item.field_name, item.value, item.pdf_label, item.page_number or "", item.confidence, item.review_required])
    ws.append([])
    ws.append(["Submitted ID", mask_id(submitted_id)])
    ws.append(["Document ID Check", verification])
    ws.append(["Review Note", verification_note])
    ws.append(["Extraction Method", method])
    ws.append(["Source", source_name])
    ws.append(["Created At", datetime.now().astimezone().isoformat(timespec="seconds")])

    raw = wb.create_sheet("Source Text")
    raw.append(["Page", "Text read from authorised preview"])
    for page_number, text in pages:
        raw.append([page_number, text])

    meta = wb.create_sheet("Field Mapping")
    meta.append(["Field", "Labels searched", "Destination field (confirm later)"])
    for field, labels in DEFAULT_MAPPINGS:
        meta.append([field, ", ".join(labels), ""])

    fill = PatternFill("solid", fgColor="F47920")
    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
        sheet.freeze_panes = "A2"
        for column in sheet.columns:
            width = min(70, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
            sheet.column_dimensions[column[0].column_letter].width = width
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(output_path)
    return output_path, verification


def extract_to_workbook(pdf_bytes, output_path, submitted_id, source_name="iDocs PDF preview"):
    pages, extracted, verification, note = read_pdf_preview(pdf_bytes, submitted_id)
    output_path, _ = write_calibration_workbook(output_path, submitted_id, source_name, pages, extracted)
    return output_path, verification, extracted, pages
